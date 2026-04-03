import json
import os
#from metadata_tool import embed_metadata

# =========================
# SAMPLE FINGERPRINT
# =========================
def sample_fingerprint():
    return {
        "fpid": "FGP-239182",
        "host": "FIN-PC01",
        "user": "Channarong",
        "timestamp": "2026-04-03T14:20",
        "hash": "fa8d23a9c"
    }

# =========================
# XLSX
# =========================
def embed_xlsx(file_path, fp):
    from openpyxl import load_workbook
    import json
    wb = load_workbook(file_path)
    wb.properties.description = json.dumps(fp)   # ✅ ใช้ตัวนี้
    wb.save(file_path)

def extract_xlsx(file_path):
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    return wb.properties.description   # ✅ ตรงกัน

# =========================
# DOCX
# =========================
def embed_docx(file_path, fp):
    from docx import Document
    doc = Document(file_path)
    doc.core_properties.comments = json.dumps(fp)
    doc.save(file_path)

def extract_docx(file_path):
    from docx import Document
    doc = Document(file_path)
    return doc.core_properties.comments

# =========================
# PDF
# =========================
def embed_pdf(file_path, fp):
    import fitz
    import json
    import os

    doc = fitz.open(file_path)

    meta = doc.metadata
    meta["subject"] = json.dumps(fp)

    doc.set_metadata(meta)

    temp_path = file_path + ".tmp"

    doc.save(temp_path)   # ✅ save ไฟล์ใหม่
    doc.close()

    os.replace(temp_path, file_path)  # ✅ replace ของเดิม

def extract_pdf(file_path):
    import fitz
    doc = fitz.open(file_path)
    return doc.metadata.get("subject")

# =========================
# PNG
# =========================
def embed_png(file_path, fp):
    from PIL import Image, PngImagePlugin
    img = Image.open(file_path)
    meta = PngImagePlugin.PngInfo()
    meta.add_text("fingerprint", json.dumps(fp))
    img.save(file_path, pnginfo=meta)

def extract_png(file_path):
    from PIL import Image
    img = Image.open(file_path)
    return img.info.get("fingerprint")

# =========================
# JPG
# =========================
def embed_jpg(file_path, fp):
    from PIL import Image
    img = Image.open(file_path)
    exif = img.getexif()
    exif[270] = json.dumps(fp)
    img.save(file_path, exif=exif)

def extract_jpg(file_path):
    from PIL import Image
    img = Image.open(file_path)
    exif = img.getexif()
    return exif.get(270)

# =========================
# MAIN HANDLER
# =========================
def embed_metadata(file_path, fp):
    ext = file_path.split(".")[-1].lower()

    if ext == "xlsx":
        embed_xlsx(file_path, fp)
    elif ext == "docx":
        embed_docx(file_path, fp)
    elif ext == "pdf":
        embed_pdf(file_path, fp)
    elif ext == "png":
        embed_png(file_path, fp)
    elif ext in ["jpg", "jpeg"]:
        embed_jpg(file_path, fp)
    else:
        print("❌ Unsupported file type:", ext)

def extract_metadata(file_path):
    ext = file_path.split(".")[-1].lower()

    if ext == "xlsx":
        return extract_xlsx(file_path)
    elif ext == "docx":
        return extract_docx(file_path)
    elif ext == "pdf":
        return extract_pdf(file_path)
    elif ext == "png":
        return extract_png(file_path)
    elif ext in ["jpg", "jpeg"]:
        return extract_jpg(file_path)
    else:
        return None

# =========================
# CLI USAGE
# =========================
if __name__ == "__main__":

    import sys

    if len(sys.argv) < 3:
        print("Usage:")
        print("  python metadata_tool.py embed <file>")
        print("  python metadata_tool.py extract <file>")
        exit()

    action = sys.argv[1]
    file_path = sys.argv[2]

    if not os.path.exists(file_path):
        print("❌ File not found")
        exit()

    if action == "embed":
        fp = sample_fingerprint()
        embed_metadata(file_path, fp)
        print("✅ Metadata embedded")

    elif action == "extract":
        raw = extract_metadata(file_path)

        if raw:
            try:
                print("=== Fingerprint ===")
                print(json.dumps(json.loads(raw), indent=2))
            except:
                print("Raw:", raw)
        else:
            print("❌ No metadata found")

    else:
        print("❌ Unknown command")
